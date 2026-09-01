"""Persistent Excel results log for completed SmartAdvisor jobs."""

from __future__ import annotations

import sys
import threading
from datetime import datetime
from pathlib import Path

from .models import SmartAdvisorJob

HEADER = (
    "Timestamp",
    "Message ID",
    "Claim ID",
    "DOS From",
    "Expected Amount",
    "Site ID",
    "Disposition",
    "Paid Amount",
    "Paid Date",
    "Check Number",
    "Denial Code",
    "Reply Sent",
)


def dataset_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent / "data"
    return Path.cwd() / "data"


def results_workbook_path() -> Path:
    return dataset_dir() / "pcc_results.xlsx"


class ResultsWorkbook:
    """Appends one row per completed job to a single persistent .xlsx file."""

    def __init__(self, path: Path | None = None) -> None:
        self.path = path or results_workbook_path()
        self._lock = threading.Lock()

    def append_result(
        self,
        job: SmartAdvisorJob,
        helper_result: dict[str, object],
        *,
        reply_sent: bool,
    ) -> None:
        from openpyxl import Workbook, load_workbook

        with self._lock:
            self.path.parent.mkdir(parents=True, exist_ok=True)
            if self.path.exists():
                workbook = load_workbook(self.path)
                sheet = workbook.active
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Results"
                sheet.append(HEADER)

            sheet.append(
                (
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    job.source_message_id,
                    job.claim_id,
                    job.dos_from,
                    job.expected_amount,
                    job.site_id,
                    str(helper_result.get("disposition") or ""),
                    str(helper_result.get("paid_amount") or ""),
                    str(helper_result.get("paid_date") or ""),
                    str(helper_result.get("check_number") or ""),
                    str(helper_result.get("denial_code") or ""),
                    "Y" if reply_sent else "N",
                )
            )
            workbook.save(self.path)

"""Label samples.xlsx by hand-written rule, and record the rule next to the answer.

This is the ground-truth pass for the sample set (SP-1.1-59). It fills in
columns D-L - the answer a correct engine should produce - and column M, the
rule that produced it. Column M is the deliverable: it is the spec that the
keyword rules and extraction patterns in concerns.json are meant to implement,
written against real inbound mail rather than against the invented examples.

Two things are deliberate:

1. A REASON is only recorded when the email states the disposition itself. The
   reasons in the taxonomy ("Not a bill on file", "Completed processing and
   denied") are outcomes an agent records AFTER looking the bill up. An inbound
   provider email asks a question; it does not announce the answer. So E is
   blank on almost every row, and that is the finding, not a gap.

2. Extraction is label-anchored only. An unlabelled number is never claimed for
   a field, because DOS/DOI/DOB share one date pattern and claim id / patient
   account / TIN all share one digit-run shape. Reporting nothing beats
   reporting the wrong date.

Where the rule as written cannot reach the correct value, the correct value is
still written into D-L (it is ground truth, not a rule dump) and column M says
which rule missed it and why. Those rows are the backlog.

    py -3.14 scripts/label_samples.py            # write samples.xlsx in place
    py -3.14 scripts/label_samples.py --check    # report only, touch nothing
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl

WORKBOOK = Path(__file__).resolve().parent.parent / "samples.xlsx"

# --------------------------------------------------------------------------
# value shapes
# --------------------------------------------------------------------------

DATE = re.compile(r"\b\d{1,2}[/.\-]\d{1,2}[/.\-]\d{2,4}\b")
CLAIM = re.compile(r"\b[A-Z]{0,3}\d{6,12}(?:-\d{1,2})?\b", re.I)
ACCT = re.compile(r"\b[A-Z]{0,3}\d{6,14}\b", re.I)
TIN = re.compile(r"\b\d{2}-\d{7}\b|\b\d{9}\b")
MONEY = re.compile(r"\b\d{1,3}(?:,\d{3})+(?:\.\d{2})?\b|\b\d+(?:\.\d{2})?\b")

# How far past a label the value is allowed to sit. Identifiers are written
# hard against their label ("Claim #: 5513778-1"), so they get a tight window;
# dates and amounts are laid out in tables and lists, so they get a loose one.
WINDOW_ID = 25
WINDOW_LIST = 60

# --------------------------------------------------------------------------
# label aliases, longest first so "claim number" claims its text before "claim"
# --------------------------------------------------------------------------

FIELDS: dict[str, dict] = {
    "claim_id": {
        "col": 6,
        "aliases": [
            "claim number", "your claim #", "claim no", "claim #", "claim#",
            "clm", "claim",
        ],
        # "Bill/Claim #" and "Ref #" are the provider's own numbering, not the
        # carrier claim id. Both appear alongside a real claim number.
        "reject_prefix": ["bill/", "our bill", "ref", "bill "],
        "value": CLAIM,
        # One email carries one claim id. Taking the first label that yields a
        # value stops a bare "claim" further down ("we submitted claim through
        # electronically") from scraping the TIN off the next line.
        "mode": "single",
    },
    "dos": {
        "col": 7,
        "aliases": ["dates of service", "date of service", "date service", "dos"],
        "reject_prefix": [],
        "value": DATE,
        "mode": "per_label",
    },
    "patient_account": {
        "col": 8,
        "aliases": [
            "patient account#", "patient account", "patient acct", "account ref#",
            "account ref", "account#", "acct#", "acc.#", "account no", "account",
            "acct",
        ],
        "reject_prefix": [],
        "value": ACCT,
        "mode": "single",
    },
    "provider_tin": {
        "col": 9,
        "aliases": ["tax id#", "tax id", "tax i.d", "taxid", "tin", "ein"],
        # "Provider ID: 330118842" is not a TIN even though it is nine digits.
        "reject_prefix": ["provider ", "prov "],
        "value": TIN,
        "mode": "single",
    },
    "expected_amount": {
        "col": 10,
        "aliases": [
            "charged amount", "charge amount", "billed amount", "billing amount",
            "appeal amount", "expected amount", "amount billed", "bill amount",
            "billed amt", "billed", "amount", "tc",
        ],
        # A payment already made is not what was billed.
        "reject_prefix": ["paid ", "check ", "amount "],
        "value": MONEY,
        # "Billed Amount : 604.20 AND 3377.40" is two real figures on one line.
        # Stopping at the newline is what keeps the next line's "Ref #: 1202143"
        # from being read as a third amount.
        "mode": "same_line",
    },
    "doi": {
        "col": 11,
        "aliases": ["date of injury", "date injury", "doi"],
        "reject_prefix": [],
        "value": DATE,
        "mode": "per_label",
    },
    "dob": {
        "col": 12,
        "aliases": ["date of birth", "date birth", "dob"],
        "reject_prefix": [],
        "value": DATE,
        "mode": "per_label",
    },
}


def _label_spans(text: str, aliases: list[str], reject_prefix: list[str]):
    """Every alias occurrence, longest alias first, no region claimed twice."""
    low = text.lower()
    spans: list[tuple[int, int]] = []
    for alias in aliases:
        for m in re.finditer(rf"(?<![a-z]){re.escape(alias)}(?![a-z])", low):
            if any(m.start() < e and m.end() > s for s, e in spans):
                continue
            before = low[max(0, m.start() - 10) : m.start()]
            if any(before.endswith(p) for p in reject_prefix):
                continue
            spans.append((m.start(), m.end()))
    return sorted(spans)


def extract(text: str, spec: dict) -> list[str]:
    """Label-anchored values, in document order, de-duplicated.

    Three modes, because the three kinds of field behave differently in real
    mail. `single` - one value per email, first label that yields wins.
    `per_label` - one value per label occurrence (three DOS lines, three dates).
    `same_line` - every value on the first line the label reaches, and no
    further (amounts are listed inline, but the next line is a different field).
    """
    mode = spec.get("mode", "per_label")
    window_len = WINDOW_ID if mode == "single" else WINDOW_LIST
    out: list[str] = []

    for _, end in _label_spans(text, spec["aliases"], spec["reject_prefix"]):
        window = text[end : end + window_len]
        # A date is never an amount and an amount is never a date, but the two
        # patterns overlap ("$08/19/2026", "9/14/2026  $812.75"). Mask the loser.
        blocked = (
            [] if spec["value"] is DATE else [m.span() for m in DATE.finditer(window)]
        )
        found: list[str] = []
        line_end = None
        for m in spec["value"].finditer(window):
            if any(m.start() < be and m.end() > bs for bs, be in blocked):
                continue
            if line_end is not None and m.start() > line_end:
                break
            v = m.group(0).strip()
            if v:
                found.append(v)
            if mode != "same_line":
                break
            if line_end is None:
                nl = window.find("\n", m.end())
                line_end = len(window) if nl == -1 else nl
        for v in found:
            if v not in out:
                out.append(v)
        if out and mode == "single":
            break
    return out


# --------------------------------------------------------------------------
# concern / reason
# --------------------------------------------------------------------------

# The sender is asking to be GIVEN a claim number - they do not have one.
CLAIM_INFO = [
    "claim number request", "provide a claim number", "provide the claim number",
    "need the claim number", "need a claim number", "what is the claim number",
    "requesting the claim",
]

# The sender is asking about a bill or claim already submitted.
BILL_STATUS = [
    "bill status", "status of the bill", "status of bill", "status of the following bill",
    "claim status", "payment status", "status update", "bill payment status",
    "provide status", "provide the status", "unpaid bill", "was received",
    "confirm if claim was received", "when can we anticipate payment",
    "status of the attached appeal", "check payment status", "claim status request",
    "eob", "check number", "date of receipt",
]

# A disposition asserted about THIS bill, in the sender's own newest words.
DISPOSITIONS = {
    "completed_processing_denied": ["completed processing and denied"],
    "not_a_bill_on_file": ["not a bill on file", "no bill on file"],
    "no_claim_on_file_missing_information": ["no claim on file", "missing information"],
    "completed_processing": ["has completed processing", "completed processing"],
}


def classify(subject: str, body: str) -> tuple[str, str, str]:
    """Returns (concern, reason, which rule fired)."""
    blob = f"{subject}\n{body}".lower()

    for phrase in CLAIM_INFO:
        if phrase in blob:
            return (
                "Claim Information",
                "Claim number request",
                f'D2: sender asks to be GIVEN a claim number - matched "{phrase}"; '
                f"E: the one reason an inbound email can state about itself",
            )

    for phrase in BILL_STATUS:
        if phrase in blob:
            for reason, cues in DISPOSITIONS.items():
                for cue in cues:
                    if cue in blob:
                        return (
                            "Bill Status",
                            reason,
                            f'D1: matched "{phrase}"; E: disposition asserted - "{cue}"',
                        )
            return (
                "Bill Status",
                "",
                f'D1: sender asks the state of an already-submitted bill - '
                f'matched "{phrase}"; E blank: an inbound question states no '
                f"disposition",
            )

    return ("Other / not a tracked concern", "", "D3: no concern phrase matched")


# --------------------------------------------------------------------------
# corrections the rule above cannot reach
# --------------------------------------------------------------------------
#
# Keyed by Email # (column A). Each entry is (column letter -> (value, note)).
# The value is ground truth. The note says why the rule missed it, and is
# appended to column M. These rows are the work list for concerns.json.

OVERRIDES: dict[int, dict[str, tuple[str | None, str]]] = {
    4: {
        "K": (None, "DOI is written with DOTS (03.11.1994). This rule accepts "
                    "[./-] so it reads; the engine's `us_date` pattern does not. "
                    "FIX concerns.json: widen us_date to accept dot separators."),
        "L": ("", "A date of birth follows the patient name with no DOB label. "
                  "Correctly refused - an unlabelled date must never be claimed, "
                  "because DOS/DOI/DOB share one pattern. Nothing to fix."),
    },
    5: {
        "J": ("", "The 'Amount:' line holds a DATE, not a figure - the sender "
                  "mis-keyed it. Correctly refused by the date/amount mask. "
                  "Nothing to fix; this row is the regression test for it."),
    },
    12: {
        "G": ("10/08/26 | 11/03/26 | 11/06/26 | 11/10/26 | 11/12/26",
              "FIVE dates of service sit under ONE 'dates of service' label, "
              "one per line, each followed by its own amount. Only the first is "
              "inside the label window. FIX: when a date label is followed by a "
              "list, read the whole list, pairing each date with the amount on "
              "its line."),
    },
    13: {
        "G": ("9/14/2026 | 9/28/2026",
              "The bills are a TABLE: a 'DOS' column header with two rows under "
              "it, and the second row is far past the label window. FIX: treat a "
              "label followed by other column headers as a column, and read the "
              "values beneath it."),
        "J": ("812.75 | 778.05",
              "Same table, 'AMOUNT BILLED' column. Same fix as G. Note the "
              "quoted call log on this row also contains 'not on file' and "
              "'denied' - see column E."),
    },
    14: {
        "H": ("", "The patient account is in the SUBJECT with no label "
                  "(a bare alphanumeric after the patient name). Correctly "
                  "refused. Only worth fixing if a subject-position rule is "
                  "wanted, and that is risky - the subject carries the claim id "
                  "in the same unlabelled shape."),
    },
}


def main() -> int:
    check_only = "--check" in sys.argv
    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb["Sheet1"]

    ws.cell(1, 13).value = "Rule / Logic"

    for r in range(2, ws.max_row + 1):
        email_no = ws.cell(r, 1).value
        subject = str(ws.cell(r, 2).value or "")
        body = str(ws.cell(r, 3).value or "")
        text = f"{subject}\n{body}"

        concern, reason, rule_note = classify(subject, body)
        ws.cell(r, 4).value = concern
        ws.cell(r, 5).value = reason

        notes = [rule_note]
        over = OVERRIDES.get(email_no, {})

        for name, spec in FIELDS.items():
            col = spec["col"]
            letter = chr(ord("A") + col - 1)
            values = extract(text, spec)
            if letter in over:
                value, why = over[letter]
                if value is None:
                    ws.cell(r, col).value = " | ".join(values) if values else None
                else:
                    ws.cell(r, col).value = value or None
                notes.append(f"{letter}: {why}")
            else:
                ws.cell(r, col).value = " | ".join(values) if values else None
                if values:
                    notes.append(
                        f"{letter}: label-anchored on "
                        + ", ".join(f'"{a}"' for a in _matched_aliases(text, spec))
                    )
                else:
                    notes.append(f"{letter}: no label present - correctly blank")

        ws.cell(r, 13).value = "\n".join(notes)
        ws.cell(r, 13).alignment = openpyxl.styles.Alignment(
            wrap_text=True, vertical="top"
        )

    ws.column_dimensions["M"].width = 70

    if check_only:
        print("check only - nothing written")
        return 0
    wb.save(WORKBOOK)
    print(f"wrote {WORKBOOK.name}")
    return 0


def _matched_aliases(text: str, spec: dict) -> list[str]:
    """Which label words actually anchored a value - this is what column M cites."""
    mode = spec.get("mode", "per_label")
    window_len = WINDOW_ID if mode == "single" else WINDOW_LIST
    low = text.lower()
    seen: list[str] = []
    for start, end in _label_spans(text, spec["aliases"], spec["reject_prefix"]):
        window = text[end : end + window_len]
        blocked = (
            [] if spec["value"] is DATE else [m.span() for m in DATE.finditer(window)]
        )
        for m in spec["value"].finditer(window):
            if any(m.start() < be and m.end() > bs for bs, be in blocked):
                continue
            alias = low[start:end]
            if alias not in seen:
                seen.append(alias)
            break
        if seen and mode == "single":
            break
    return seen


if __name__ == "__main__":
    raise SystemExit(main())

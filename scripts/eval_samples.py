"""Score the engine against the hand-labelled ground truth in samples.xlsx.

Columns D-L of samples.xlsx are what a correct engine should produce; this runs
the real `TriageEngine` over the same 15 emails and reports where it agrees.

Value comparison is normalized on both sides - the engine emits `2026-07-09`
and `2288.60` where the sheet holds whatever the labeller wrote - so the
comparison is about whether the same VALUE was found, not the same formatting.
A ground-truth cell holding several values ("11/21/2041 | 11/24/2041") is scored
as a SET against `FieldValue.values`: every distinct value must be found and no
extra one invented. Row multiplicity is validated separately through line-item
tests because the workbook's ground-truth cells store unique values only.
Partial credit is reported separately as PARTIAL, so a field that finds 2 of 5
dates is not counted the same as one that finds all 5.

    py -3.14 scripts/eval_samples.py
    py -3.14 scripts/eval_samples.py --verbose    # per-row detail
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from email_triage.engine import TriageEngine  # noqa: E402
from email_triage.extract import normalize_value  # noqa: E402

WORKBOOK = Path(__file__).resolve().parent.parent / "samples.xlsx"

# sheet column -> (engine field name, normalizer used to compare)
COLUMNS = {
    6: ("claim_id", "upper_alnum"),
    7: ("date_of_service", "date_iso"),
    8: ("patient_account", "upper_alnum"),
    9: ("provider_tin", "digits"),
    10: ("expected_amount", "money"),
    11: ("date_of_injury", "date_iso"),
    12: ("date_of_birth", "date_iso"),
}


def truth_values(cell, normalizer: str) -> set[str]:
    if cell is None or str(cell).strip() == "":
        return set()
    return {
        normalize_value(part.strip(), normalizer)
        for part in str(cell).split("|")
        if part.strip()
    }


def main() -> int:
    verbose = "--verbose" in sys.argv
    engine = TriageEngine()
    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb["Sheet1"]

    concern_ok = 0
    rows = 0
    per_field = {name: [0, 0, 0, 0] for name, _ in COLUMNS.values()}  # hit, miss, wrong, partial
    review = 0

    print(f"engine {'3 layers' if engine.embeddings_active else '2 layers (no model)'}"
          f"  |  config {engine.config.config_version}\n")

    for r in range(2, ws.max_row + 1):
        rows += 1
        subject = str(ws.cell(r, 2).value or "")
        body = str(ws.cell(r, 3).value or "")
        want_concern = str(ws.cell(r, 4).value or "")

        res = engine.classify(body, subject=subject)
        got_concern = res.display_name or ""
        match = got_concern == want_concern
        concern_ok += match
        if res.needs_review:
            review += 1

        detail = []
        for col, (fname, norm) in COLUMNS.items():
            want = truth_values(ws.cell(r, col).value, norm)
            fv = res.fields.get(fname)
            got = fv.value if fv else None
            if not want:
                # Nothing expected. Finding anything is a false positive.
                if got:
                    per_field[fname][2] += 1
                    detail.append(f"{fname}=FALSE-POSITIVE")
                continue
            found = set(fv.values) if fv and fv.values else ({got} if got else set())
            if found == want:
                per_field[fname][0] += 1
            elif found and found < want:
                per_field[fname][3] += 1
                detail.append(f"{fname}=PARTIAL({len(found)}/{len(want)})")
            elif found:
                per_field[fname][2] += 1
                detail.append(f"{fname}=WRONG")
            else:
                per_field[fname][1] += 1
                detail.append(f"{fname}=MISSED")

        flag = "ok " if match else "BAD"
        line = (
            f"  #{ws.cell(r, 1).value:>2}  {flag}  {got_concern or '(none)':<20}"
            f"{res.status.value:<14} {res.confidence:>4.0%}"
            f"{'  REVIEW' if res.needs_review else ''}"
        )
        if verbose and detail:
            line += "   " + " ".join(detail)
        print(line)

    print(f"\nconcern      : {concern_ok}/{rows}")
    print(f"needs review : {review}/{rows}")
    print(
        f"\n  {'field':<18} {'exact':>6} {'partial':>8}"
        f" {'missed':>7} {'wrong':>6}  of truth"
    )
    for name, (hit, miss, wrong, partial) in per_field.items():
        total = hit + partial + miss + wrong
        print(
            f"  {name:<18} {hit:>6} {partial:>8}"
            f" {miss:>7} {wrong:>6}  {total:>3}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
